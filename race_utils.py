"""
Shared utility functions for Pinewood Derby heat generation and result processing.
This module is intended to be imported into both heats.py and results.py
"""

import math
import os
import re
import secrets
import zipfile
import random

from collections import defaultdict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def is_nan(val):
    """
    Check if a value is a NaN (Not a Number).

    Parameters:
        val (any): The value to check.

    Returns:
        bool: True if val is a float NaN, False otherwise.
    """
    return isinstance(val, float) and math.isnan(val)


def secure_shuffle(lst):
    """
    Securely shuffle a list using cryptographically strong randomness.

    Parameters:
        lst (list): The list to shuffle.

    Returns:
        list: A new shuffled version of the list.
    """
    lst_copy = lst[:]
    for i in reversed(range(1, len(lst_copy))):
        j = secrets.randbelow(i + 1)
        lst_copy[i], lst_copy[j] = lst_copy[j], lst_copy[i]
    return lst_copy


def sanitize_sheet_title(title):
    """
    Clean and truncate an Excel sheet name to comply with Excel's constraints.

    Parameters:
        title (str): Original sheet title.

    Returns:
        str: Sanitized title (max 31 characters, safe characters only).
    """
    title = re.sub(r"[^A-Za-z0-9 _-]", "", title)
    title = title.replace(" ", "_")
    return title[:31]


def _validate_excel_file(filename):
    """
    Validates that the provided filename is an existing, valid Excel .xlsx file.

    Parameters:
    ----------
    filename : str
        Path to the file to validate.

    Raises:
    -------
    FileNotFoundError: If the file does not exist.
    ValueError: If the file is not an .xlsx file or not a valid zip file.
    """
    if not os.path.isfile(filename):
        raise FileNotFoundError(f"Excel file not found: {filename}")
    if not filename.lower().endswith(".xlsx"):
        raise ValueError(f"Invalid file format (must be .xlsx): {filename}")
    if not zipfile.is_zipfile(filename):
        raise ValueError(f"File is not a valid Excel zip file: {filename}")

def read_excel_sheet(filename, sheet_name=None):
    """
    Safely read an Excel sheet into a DataFrame with validation.

    Parameters:
        filename (str): Path to the Excel file.
        sheet_name (str|None): Specific sheet name to read (default: first sheet).

    Returns:
        pd.DataFrame: Contents of the sheet.

    Raises:
        ValueError: If file is not valid or reading fails.
    """
    _validate_excel_file(filename)

    try:
        return pd.read_excel(filename, sheet_name=sheet_name)
    except Exception as e:
        raise ValueError(f"Failed to read sheet '{sheet_name}': {e}")


def get_excel_sheet_names(filename):
    """
    Validates the file and retrieves sheet names from an Excel workbook.

    Args:
        filename (str): Path to the Excel file.

    Returns:
        list[str]: List of sheet names if successful.

    Raises:
        FileNotFoundError: If the file doesn't exist or is not a .xlsx file.
        ValueError: If the Excel file contains no sheets.
        Exception: If the file can't be opened as an Excel file.
    """
    _validate_excel_file(filename)

    try:
        sheet_names = pd.ExcelFile(filename).sheet_names
        if not sheet_names:
            raise ValueError("Excel file contains no sheets.")
        return sheet_names
    except Exception as e:
        raise Exception(f"Could not open file '{filename}': {e}")


def format_all_sheets(filename):
    """
    Formats all sheets in the given Excel workbook:
    - Center-aligns all cells.
    - Uses fixed-width font (Courier New).
    - Bolds header rows.
    - Auto-adjusts column widths based on content.
    - Applies light gray background to even-numbered heats (if applicable).

    Parameters:
        filename (str): Path to the Excel workbook
    """
    book = load_workbook(filename)

    light_gray_fill = PatternFill(
        start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
    )

    for sheet_name in book.sheetnames:
        sheet = book[sheet_name]

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Determine if we should apply even-heat highlighting
        apply_even_heat_fill = sheet_name not in (
            "Racers",
            "Runoff",
        ) and not sheet_name.endswith("_Rankings")

        # Identify the Heat column index
        heat_col_index = None
        if apply_even_heat_fill:
            for col_idx, cell in enumerate(sheet[1]):
                if str(cell.value).strip().lower() == "heat":
                    heat_col_index = col_idx
                    break

        # Bold headers
        for cell in sheet[1]:
            cell.font = Font(bold=True, name="Courier New")
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Format all other cells
        for row in sheet.iter_rows(min_row=2):
            is_even_heat = False
            if apply_even_heat_fill and heat_col_index is not None:
                try:
                    heat_val = int(row[heat_col_index].value)
                    is_even_heat = heat_val % 2 == 0
                except (TypeError, ValueError):
                    pass

            for cell in row:
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(name="Courier New")
                if is_even_heat:
                    cell.fill = light_gray_fill

        # Adjust column widths with a scaling factor to prevent crunched headers
        for column_cells in sheet.columns:
            header = column_cells[0]
            max_length = len(str(header.value or ""))

            for cell in column_cells[1:]:
                try:
                    max_length = max(max_length, len(str(cell.value or "")))
                except Exception:
                    pass

            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[header.column_letter].width = adjusted_width

    book.save(filename)
    print("[OK] All sheets formatted successfully.")


def get_racer_heats(filename):
    """
    Extracts heat participation data for each car from an Excel workbook containing
    race schedules.

    This function reads all relevant sheets in the workbook and builds a dictionary
    that maps each car to the list of heat numbers it is scheduled to race in.
    It uses the actual values from the "Heat" column (not row numbers), ensuring
    accurate tracking of car participation across heats.

    Parameters:
    ----------
    filename : str
        Path to the Excel workbook that contains heat data.

    Returns:
    -------
    dict
        A dictionary mapping car identifiers (as strings) to a list of heat numbers (integers).
        Example:
        {
            "101": [1, 3, 7],
            "202": [2, 5, 6],
            ...
        }

    Behavior:
    ---------
    - Skips sheets named "Racers", "Runoff", or any that end in "_Rankings".
    - Expects each heat sheet to contain at least two columns: "Car" and "Heat".
    - Each row represents one entry of a car in a specific heat and lane.

    Data Handling:
    --------------
    - Car values are cast to strings for consistency.
    - Heat values are cast to integers and must not be NaN.
    - If a car appears in multiple sheets, its heat numbers are aggregated.

    Error Handling:
    ---------------
    - Any sheet that cannot be read due to formatting or I/O errors is skipped
      with a warning message.
    - Sheets missing required columns are silently skipped.

    Dependencies:
    -------------
    - `get_excel_sheet_names(filename)`: Returns a list of sheet names in the Excel file.
    - `read_excel_sheet(filename, sheet_name)`: Reads a specified sheet into a DataFrame.

    Example:
    --------
    >>> get_racer_heats(\"heats.xlsx\")
    {
        \"103\": [1, 4, 7],
        \"204\": [2, 3],
        ...
    }
    """
    heats_data = {}

    sheet_names = get_excel_sheet_names(filename)

    for sheet_name in sheet_names:
        if sheet_name in ("Racers", "Runoff") or sheet_name.endswith("_Rankings"):
            continue

        try:
            df_heat = read_excel_sheet(filename, sheet_name=sheet_name)
        except Exception as e:
            print(f"[WARN] Skipping invalid sheet '{sheet_name}': {e}")
            continue

        if "Car" not in df_heat.columns or "Heat" not in df_heat.columns:
            continue

        df_heat["Car"] = df_heat["Car"].astype(str)

        for _, row in df_heat.iterrows():
            car = row["Car"]
            heat = row["Heat"]
            if pd.isna(heat):
                continue

            heat = int(heat)
            if car not in heats_data:
                heats_data[car] = []
            heats_data[car].append(heat)

    return heats_data


def update_racer_heats(filename, heats_data):
    """
    Updates the "Racers" sheet in an Excel workbook by adding a summary of heat
    assignments for each car.

    This function takes a mapping of car numbers to their scheduled heat numbers
    and writes that information to a new "Heats" column in the "Racers" sheet.
    If the sheet already exists, it will be replaced with the updated version.

    Parameters:
    ----------
    filename : str
        Path to the Excel workbook containing the "Racers" sheet.

    heats_data : dict
        Dictionary mapping car identifiers (as strings or integers) to a list of
        heat numbers (integers). Example:
        {
            "101": [1, 3, 5],
            "202": [2, 4],
            ...
        }

    Behavior:
    ---------
    - Loads the "Racers" sheet and converts all car identifiers to strings.
    - Creates a new "Heats" column that lists each car's heat numbers as a
      comma-separated string (e.g., "1, 3, 5").
    - Writes the updated DataFrame back to the "Racers" sheet in the same file.
      If the sheet already exists, it is replaced.

    Error Handling:
    ---------------
    - If the file or sheet cannot be read or written, an error message is printed.
    - Failures do not raise exceptions; they log an error to the console instead.

    Dependencies:
    -------------
    - `read_excel_sheet(filename, sheet_name)`: Reads a specific sheet into a DataFrame.
    - Uses `pandas.ExcelWriter` with `openpyxl` engine for writing.

    Example:
    --------
    >>> update_racer_heats("event_schedule.xlsx", {"101": [1, 3], "102": [2, 4]})
    [OK] Racers sheet updated with heat allocations.

    Notes:
    ------
    - The function assumes that the "Racers" sheet contains a "Car" column.
    - Cars not present in `heats_data` will have an empty string in the "Heats" column.
    """
    try:
        df_racers = read_excel_sheet(filename, sheet_name="Racers")
        df_racers["Car"] = df_racers["Car"].astype(str)

        df_racers["Heats"] = df_racers["Car"].map(
            lambda num: ", ".join(map(str, heats_data.get(num, [])))
        )

        with pd.ExcelWriter(
            filename, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            df_racers.to_excel(writer, sheet_name="Racers", index=False)

        print("[OK] Racers sheet updated with heat allocations.")

    except Exception as e:
        print(f"[ERROR] Could not update heat allocations on 'Racers' sheet: {e}")


def analyze_opponents(heats_df):
    """
    Analyzes car matchups across all heats and returns, for each car,
    the set of unique opponents it has raced against.

    Parameters:
    ----------
    heats_df : pd.DataFrame
        DataFrame representing heat assignments with the following columns:
        - "Heat": Heat number
        - "Car": Car identifier
        - "Lane": Lane label (not used in this function)

    Returns:
    -------
    dict
        A dictionary where each key is a car, and the value is a set of other cars
        (opponents) that the car has shared a heat with.

    Example:
    --------
    Input heats:
        Heat | Car
        -----|-----
         1   | A
         1   | B
         1   | C
         2   | A
         2   | D

    Output:
        {
            "A": {"B", "C", "D"},
            "B": {"A", "C"},
            "C": {"A", "B"},
            "D": {"A"}
        }

    Notes:
    ------
    This function is essential for evaluating opponent diversity and fairness.
    The output is typically passed to `opponent_fairness_score()` for scoring.
    """
    opponents = defaultdict(set)
    for _, heat in heats_df.groupby("Heat"):
        cars = set(heat["Car"])
        for car in cars:
            opponents[car].update(cars - {car})
    return opponents


def opponent_fairness_score(opponents):
    """
    Computes a fairness score based on how evenly opponents are distributed
    across all cars.

    The idea is that a fair schedule should give each car a similar number
    of unique opponents. This function penalizes variance in those counts.

    Parameters:
    ----------
    opponents : dict
        A dictionary where keys are car identifiers and values are sets of
        opponents (as produced by `analyze_opponents()`).

    Returns:
    -------
    float
        A numeric fairness score. Lower values indicate better (more even)
        opponent distribution.

    Calculation:
    ------------
    - For each car, count the number of unique opponents.
    - Compute the mean of those counts.
    - Return the variance (mean squared deviation from the average).

    Example:
    --------
    Opponents = {
        "A": {"B", "C"},
        "B": {"A", "C"},
        "C": {"A", "B"},
        "D": {"E"},
        "E": {"D"}
    }

    A, B, C have 2 opponents; D, E have 1. Average = 1.6
    Variance ≈ 0.24 → fairness score = 0.24

    Notes:
    ------
    - This metric assumes that each car should have as close to the same number
      of unique opponents as possible.
    - Can be used to compare different heat assignment strategies.
    """
    counts = [len(opp) for opp in opponents.values()]
    avg = sum(counts) / len(counts)
    return sum((count - avg) ** 2 for count in counts) / len(counts)


def optimize_opponent_fairness(heats_df, iterations=500):
    """
    Attempts to optimize fairness in car matchups across heats by minimizing
    repeated opponents, while maintaining valid lane assignments and ensuring
    no duplicate cars within the same heat ("Perfect-N" constraints).

    Purpose:
    --------
    This function improves competitive fairness by ensuring that each car races
    against a diverse set of opponents, rather than repeatedly racing the same
    few cars. It does this by randomly swapping cars between heats (in the same
    lane) and evaluating whether the overall fairness score improves.

    Parameters:
    ----------
    heats_df : pd.DataFrame
        The initial set of heat assignments with columns:
        - "Heat": Heat number
        - "Car": Car identifier
        - "Lane": Lane label

    iterations : int, optional (default=500)
        Number of random trial swaps to attempt. More iterations can lead to better
        fairness, at the cost of performance.

    Returns:
    -------
    pd.DataFrame
        A new DataFrame with optimized heat assignments based on opponent fairness.

    Constraints (Perfect-N Preserved):
    ----------------------------------
    - Cars are only swapped within the same lane across different heats.
    - Swaps are only applied if they do not introduce a duplicate car within any heat.
    - Lane uniqueness and heat structure remain unchanged.

    Process:
    --------
    1. Measure initial fairness using `analyze_opponents()` + `opponent_fairness_score()`.
    2. Repeat for `iterations`:
        - Randomly select two heats and a lane.
        - If both heats have cars in that lane, propose a swap.
        - Validate the swap does not cause duplicate cars in either heat.
        - If the swap improves the fairness score, keep it as the new best.
    3. After all iterations, return the best version found.

    Logs:
    -----
    Prints a summary showing:
    - Initial opponent fairness score
    - Final (optimized) score
    - Total improvement achieved

    Example:
    --------
    >>> optimized_df = optimize_opponent_fairness(heats_df, iterations=1000)
    Fairness optimization completed: initial score = 2.350, optimized score = 1.825, improvement = 0.525

    Notes:
    ------
    - `analyze_opponents()` should produce a mapping of which cars raced each other.
    - `opponent_fairness_score()` should compute a numeric score (lower is fairer).
    - This method is probabilistic and may yield different results on each run.
    - Works best after `rebalance_heats()` to ensure reasonably full heats.
    """
    best_df = heats_df.copy()
    initial_score = opponent_fairness_score(analyze_opponents(best_df))
    best_score = initial_score

    lanes = heats_df["Lane"].unique()
    heats = heats_df["Heat"].unique()

    for _ in range(iterations):
        trial_df = best_df.copy()

        # Select two random heats and one random lane
        h1, h2 = random.sample(list(heats), 2)
        lane = random.choice(lanes)

        # Swap the cars in the selected lane
        idx1 = trial_df.index[(trial_df["Heat"] == h1) & (trial_df["Lane"] == lane)]
        idx2 = trial_df.index[(trial_df["Heat"] == h2) & (trial_df["Lane"] == lane)]

        if not idx1.empty and not idx2.empty:
            car1 = trial_df.at[idx1[0], "Car"]
            car2 = trial_df.at[idx2[0], "Car"]

            # Check if swapping introduces duplicate cars
            heat1_cars = set(trial_df.loc[trial_df["Heat"] == h1, "Car"])
            heat2_cars = set(trial_df.loc[trial_df["Heat"] == h2, "Car"])

            if car2 not in heat1_cars and car1 not in heat2_cars:
                trial_df.at[idx1[0], "Car"] = car2
                trial_df.at[idx2[0], "Car"] = car1

        trial_score = opponent_fairness_score(analyze_opponents(trial_df))

        if trial_score < best_score:
            best_df, best_score = trial_df, trial_score

    improvement = initial_score - best_score
    print(
        f"Fairness optimization completed: initial score = {initial_score:.3f}, optimized score = {best_score:.3f}, improvement = {improvement:.3f}"
    )

    return best_df


def rebalance_heats(heats_df, num_lanes):
    """
    Attempts to rebalance underfilled heats by shifting car-lane assignments
    from fully populated heats, while preserving lane and car uniqueness
    within each heat ("Perfect-N" constraints).

    Purpose:
    --------
    In some randomized heat generation scenarios, a heat may end up with exactly
    (num_lanes - 2) cars — not technically invalid, but less desirable.
    This function seeks to improve overall balance by "donating" a valid car-lane
    entry from a fully populated heat into the underfilled one, bumping it up to
    (num_lanes - 1) cars where possible.

    Parameters:
    ----------
    heats_df : pd.DataFrame
        DataFrame containing heat assignments with the following required columns:
        - "Heat": Heat number
        - "Car": Car identifier
        - "Lane": Lane label

    num_lanes : int
        Total number of lanes available per heat. This defines what "fully occupied"
        and "underfilled" means in the context of the algorithm.

    Returns:
    -------
    pd.DataFrame
        Updated and re-sorted DataFrame with improved heat balance where possible.

    Constraints (Perfect-N Rules):
    ------------------------------
    - A car must not appear more than once in the same heat.
    - A lane must not be assigned to more than one car in the same heat.

    How It Works:
    -------------
    - Iterates through all heats looking for those with exactly (num_lanes - 2) cars.
    - For each such heat, searches other fully populated heats (exactly `num_lanes` cars)
      for a car-lane pair that can be safely moved without violating constraints.
    - If a valid candidate is found, it is reassigned to the underfilled heat.
    - Repeats this process for all underfilled heats, one move at a time.

    Logs:
    -----
    Prints a summary indicating how many underfilled heats were detected and how many
    remained after the optimization attempt.

    Example:
    --------
    >>> rebalance_heats(heats_df, num_lanes=4)
    Optimization complete: Unbalanced heats before: 2, after: 0

    Notes:
    ------
    - This optimization is conservative: only one move is attempted per underfilled heat.
    - Designed for use after initial heat generation but before fairness optimization.
    """
    before_unbalanced = heats_df["Heat"].value_counts().eq(num_lanes - 2).sum()

    heat_numbers = heats_df["Heat"].unique()

    for heat_num in heat_numbers:
        current_heat = heats_df[heats_df["Heat"] == heat_num]

        # Only rebalance heats with exactly (num_lanes - 2) cars
        if len(current_heat) != num_lanes - 2:
            continue

        used_cars = set(current_heat["Car"])
        used_lanes = set(current_heat["Lane"])

        # Search fully occupied heats for valid moves
        for donor_heat_num in heat_numbers:
            if donor_heat_num == heat_num:
                continue

            donor_heat = heats_df[heats_df["Heat"] == donor_heat_num]

            # Skip heats that aren't fully occupied
            if len(donor_heat) != num_lanes:
                continue

            # Look for a valid car/lane to move
            for idx, row in donor_heat.iterrows():
                car, lane = row["Car"], row["Lane"]

                # Check Perfect-N constraints
                if car not in used_cars and lane not in used_lanes:
                    # Move the car-lane pair to the current heat
                    heats_df.at[idx, "Heat"] = heat_num

                    # Resort dataframe
                    heats_df = heats_df.sort_values(
                        by=["Heat", "Lane", "Car"]
                    ).reset_index(drop=True)
                    break  # Move to the next heat after a successful rebalance
            else:
                continue  # Continue if inner loop wasn't broken
            break  # Break if inner loop was successful

    after_unbalanced = heats_df["Heat"].value_counts().eq(num_lanes - 2).sum()
    print(
        f"Optimization complete: Unbalanced heats before: {before_unbalanced}, after: {after_unbalanced}"
    )

    return heats_df
