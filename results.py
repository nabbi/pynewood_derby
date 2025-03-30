#!/usr/bin/env python3
"""
results.py — Pinewood Derby Results Processor

Processes heat result sheets from an Excel workbook and calculates standings
based on total points, 1st-place finishes, and heats run. Updates the Racers
sheet with summary stats and writes ranking sheets per heat group.

Usage:
    python results.py <results_file.xlsx>
"""

import sys
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from scipy.stats import rankdata

from race_utils import format_all_sheets, get_excel_sheet_names, read_excel_sheet


def calculate_opponent_uniqueness(df):
    """
    Calculates the percentage of unique opponents each car has raced against,
    relative to the total number of possible opponents in the dataset.

    This metric is useful for evaluating fairness and diversity in race
    scheduling—higher percentages indicate a car has been matched against a
    broader range of competitors.

    Parameters:
    ----------
    df : pd.DataFrame
        A DataFrame containing at least the following columns:
        - "Car": Car identifier
        - "Heat": Heat number each car participated in

    Returns:
    -------
    pd.DataFrame
        A DataFrame with the following columns:
        - "Car": Car identifier (as a string)
        - "Opponent_Uniqueness_Pct": Percentage (0–100) representing how many unique
          competitors this car has faced, relative to the total possible opponents.

        Example:
        ----------
        +------+--------------------------+
        | Car  | Opponent_Uniqueness_Pct |
        +------+--------------------------+
        | 101  | 87.5                     |
        | 102  | 62.5                     |
        +------+--------------------------+

    Calculation:
    ------------
    For each car:
        - Find all heats the car appears in.
        - Collect all other cars that appear in those heats.
        - Compare that set to the total possible opponents (all cars minus itself).
        - Compute the percentage of unique opponents faced.

    Notes:
    ------
    - Car identifiers are cast to strings to ensure consistency.
    - If a car has no possible opponents (i.e. it’s the only car), it receives 0%.
    - Ideal for use in fairness audits or schedule diagnostics after heat generation.

    Example:
    --------
    >>> calculate_opponent_uniqueness(df)
       Car  Opponent_Uniqueness_Pct
    0  101                     85.7
    1  102                     92.9
    ...
    """
    car_opponents = {}
    all_cars = set(df["Car"].astype(str).unique())
    df["Car"] = df["Car"].astype(str)

    for car in all_cars:
        heats = df[df["Car"] == car]["Heat"].unique()
        opponents = set()
        for heat in heats:
            competitors = set(df[df["Heat"] == heat]["Car"])
            competitors.discard(car)
            opponents.update(competitors)

        total_possible_opponents = all_cars - {car}
        percentage = (
            (len(opponents) / len(total_possible_opponents)) * 100
            if total_possible_opponents
            else 0.0
        )
        car_opponents[car] = round(percentage, 1)

    return pd.DataFrame(
        list(car_opponents.items()), columns=["Car", "Opponent_Uniqueness_Pct"]
    )


def update_racers_tab(path, summary_data):
    """
    Updates the 'Racers' sheet in the Excel workbook with performance summary data,
    including total points, heat count, first-place finishes, and overall rank.

    This function merges computed race summary statistics into the main "Racers" sheet,
    ensuring all relevant results are attached to each car, and then sorts the data for
    readability and reporting.

    Parameters:
    ----------
    path : str
        The full file path to the Excel workbook.

    summary_data : pd.DataFrame
        A DataFrame containing summarized race results for each car.
        Expected columns:
            - "Car" (string/int): Unique car identifier
            - "Total_Points" (numeric): Sum of points earned by the car
            - "First_Place_Count" (int): Number of first-place finishes
            - "Rank" (int): Final ranking based on performance

    Behavior:
    ---------
    - Reads the existing "Racers" sheet and ensures car identifiers are treated as strings.
    - Drops any existing columns that may conflict with the new summary data:
        - "Total_Points", "First_Place_Count", "Rank"
    - Merges the summary data into the sheet by matching on "Car".
    - Sorts the resulting sheet by "Class", "Group", and "Rank" for organized viewing.
    - Rewrites the updated DataFrame back to the "Racers" sheet, replacing the original.

    Error Handling:
    ---------------
    - If any step fails (reading, merging, writing), an error message is printed.
    - No exception is raised; failure is logged for diagnostics.

    Dependencies:
    -------------
    - `read_excel_sheet(path, sheet_name)`: Reads a sheet into a DataFrame.
    - Uses `pandas.ExcelWriter` with `openpyxl` to overwrite the sheet in-place.

    Example:
    --------
    >>> update_racers_tab("results.xlsx", summary_df)
    [OK] Racers sheet updated and sorted by Class, Group, Rank.

    Notes:
    ------
    - Assumes the "Racers" sheet includes at least a "Car" column, and optionally
      "Class" and "Group" for sorting.
    - This is typically called after final standings are computed across all heats.
    """
    try:
        df_racers = read_excel_sheet(path, sheet_name="Racers")
        df_racers["Car"] = df_racers["Car"].astype(str)
        summary_data["Car"] = summary_data["Car"].astype(str)

        df_racers = df_racers.drop(
            columns=[
                col
                for col in ["Total_Points", "First_Place_Count", "Rank"]
                if col in df_racers.columns
            ]
        )

        summary_selected = summary_data[
            ["Car", "Total_Points", "First_Place_Count", "Rank"]
        ]
        merged = df_racers.merge(summary_selected, how="left", on="Car")
        merged = merged.sort_values(by=["Class", "Group", "Rank"])

        with pd.ExcelWriter(
            path, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            merged.to_excel(writer, sheet_name="Racers", index=False)

        print("[OK] Racers sheet updated and sorted by Class, Group, Rank.")

    except BaseException as err:
        print(f"[ERROR] Could not update 'Racers' sheet: {err}")


def process_results(filepath):
    """
    Processes all heat sheets in a race workbook to calculate standings,
    rank competitors, and update summary information across the file.

    This function serves as the central results pipeline. It reads each group’s
    heat data, computes total points and other stats for each car, detects tie
    scenarios, writes new ranking sheets per group, and updates the main
    "Racers" tab with combined results.

    Parameters:
    ----------
    filepath : str
        Full path to the Excel file containing race heat results and a "Racers" sheet.

    Behavior:
    ---------
    - Iterates through all sheets in the workbook, skipping special-purpose sheets:
        - "Racers", "Runoff", and any ending in "_Rankings"
    - For each valid heat sheet:
        - Filters out rows with missing or non-numeric placement values
        - Aggregates key stats per car:
            - Total Points (sum of places)
            - First Place Count
            - Heat Count
            - Opponent Uniqueness %
            - Average and Percent Heat Size
        - Sorts and ranks competitors using a competition-style ranking system
          (e.g., 1, 2, 2, 4...)
        - Detects ties in the top 3 and logs them
        - Writes rankings to a new sheet named "{Group}_Rankings"
    - Merges all group summaries into a single DataFrame and updates the "Racers" tab.

    Outputs:
    --------
    - One new sheet per group with rankings and summary stats
    - Updated "Racers" sheet with total points, firsts, and overall rank
    - Console logs for ties, warnings, and successful operations

    Dependencies:
    -------------
    - `read_excel_sheet()`, `get_excel_sheet_names()`
    - `calculate_opponent_uniqueness()`
    - `update_racers_tab()`
    - `scipy.stats.rankdata` (for competition-style rankings)

    Error Handling:
    ---------------
    - Gracefully skips unreadable or invalid sheets with warnings
    - Handles I/O and merge failures without crashing the program
    - Logs detailed messages for traceability

    Example:
    --------
    >>> process_results("pinewood_2024_results.xlsx")
    [OK] Rankings written to sheet: Tigers_Rankings
    [TIE] Tie detected in top 3 of 'Bears'
    [OK] Racers sheet updated and sorted by Class, Group, Rank.
    [INFO] Ties detected in groups: ['Bears']

    Notes:
    ------
    - Uses competition-style ranking (ties share a rank, next rank is skipped)
    - Assumes each heat sheet contains at least "Car", "Heat", and "Place" columns
    - Should be called after all heats are recorded and finalized
    """
    try:
        sheet_names = get_excel_sheet_names(filepath)
    except BaseException as err:
        print(f"[ERROR] {err}")
        return

    all_summaries = []
    tie_groups = []

    for sheet in sheet_names:
        if sheet in ("Racers", "Runoff") or sheet.endswith("_Rankings"):
            continue

        try:
            df = read_excel_sheet(filepath, sheet_name=sheet)
        except BaseException as err:
            print(f"[WARN] Could not read sheet '{sheet}': {type(err).__name__}: {err}")
            continue

        if not isinstance(df, pd.DataFrame) or df.empty or "Place" not in df.columns:
            print(f"[WARN] Sheet '{sheet}' is empty or missing 'Place'. Skipping.")
            continue

        df = df[pd.to_numeric(df["Place"], errors="coerce").notna()]
        df["Place"] = df["Place"].astype(int)

        if df.empty:
            print(f"[WARN] Sheet '{sheet}' has no usable placements. Skipping.")
            continue

        lane_count_max = df["Lane"].nunique()
        df["Heat_Size"] = df.groupby("Heat")["Car"].transform("count")
        opponent_pct_df = calculate_opponent_uniqueness(df)

        summary = df.groupby("Car", as_index=False).agg(
            Name=("Name", "first"),
            Total_Points=("Place", "sum"),
            First_Place_Count=("Place", lambda x: (x == 1).sum()),
            Heat_Count=("Place", "count"),
            Avg_Heat_Size=("Heat_Size", "mean"),
        )

        summary["Avg_Heat_Size"] = summary["Avg_Heat_Size"].round(1)
        summary["Heat_Size_Pct"] = (
            ((summary["Avg_Heat_Size"] / lane_count_max) * 100).clip(upper=100).round(1)
        )

        summary = summary.merge(opponent_pct_df, on="Car", how="left")
        summary = summary.sort_values(
            by=["Total_Points", "First_Place_Count"], ascending=[True, False]
        ).reset_index(drop=True)

        rank_array = np.array(
            list(zip(summary["Total_Points"], -summary["First_Place_Count"])),
            dtype=[("points", int), ("firsts", int)],
        )
        summary["Rank"] = rankdata(rank_array, method="min").astype(int)

        top3 = summary[summary["Rank"] <= 3]
        if top3["Rank"].duplicated().any():
            tie_groups.append(sheet)
            print(f"[TIE] Tie detected in top 3 of '{sheet}'")

        results_sheet = f"{sheet}_Rankings"
        try:
            with pd.ExcelWriter(
                filepath, engine="openpyxl", mode="a", if_sheet_exists="replace"
            ) as writer:
                summary.to_excel(writer, sheet_name=results_sheet, index=False)
            print(f"[OK] Rankings written to sheet: {results_sheet}")
        except BaseException as err:
            print(f"[ERROR] Could not write rankings for '{sheet}': {err}")

        all_summaries.append(summary)

    if all_summaries:
        try:
            full_summary = pd.concat(all_summaries, ignore_index=True)
            update_racers_tab(filepath, full_summary)
        except BaseException as err:
            print(f"[ERROR] Failed to update Racers tab: {type(err).__name__}: {err}")

    print(
        f"[INFO] Ties detected in groups: {tie_groups}"
        if tie_groups
        else "[INFO] No ties found in top 3 standings."
    )


def add_runoff_tab(filepath):
    """
    Adds a "Runoff" sheet to the Excel workbook if any group contains a tie
    within the top 3 rankings.

    This function scans all *_Rankings sheets for tied placements in the top 3
    (ranks 1, 2, or 3). If any ties are found, it creates a new "Runoff" sheet
    listing those cars, along with their identifying information, to be used
    for organizing tiebreaker races.

    Parameters:
    ----------
    filepath : str
        Path to the Excel workbook containing heat results, rankings, and racer info.

    Behavior:
    ---------
    - Loads all sheets from the workbook into memory.
    - Checks if a "Racers" sheet exists; if not, exits early.
    - Iterates through all sheets ending in "_Rankings".
        - For each, checks the "Rank" column for duplicates among the top 3.
        - If duplicates are found, merges in metadata from the "Racers" sheet
          and stores a row for each tied racer.
    - If any ties are found:
        - Creates or replaces a sheet called "Runoff".
        - Writes all qualifying rows to it, sorted by Class, Group, Rank, Name.

    Returns:
    --------
    None

    Output:
    -------
    - A "Runoff" tab containing cars that need tiebreaker races.
    - Console messages indicating success, failure, or if no action is needed.

    Sheet Columns:
    --------------
    The "Runoff" sheet will contain the following columns:
    - Car
    - Name
    - Class
    - Group
    - Description
    - LastRank (the tied rank from the original ranking sheet)

    Example:
    --------
    >>> add_runoff_tab("pinewood_final_results.xlsx")
    [OK] Runoff tab created with 6 entries.

    Dependencies:
    -------------
    - `read_excel_sheet()` — loads all relevant sheets as a dict of DataFrames.
    - `dataframe_to_rows()` — converts DataFrame to row data for Excel writing.
    - `openpyxl.load_workbook()` — used for editing the Excel workbook directly.

    Notes:
    ------
    - If the "Runoff" sheet already exists, it will be replaced.
    - If no top-3 ties are found in any ranking sheet, no tab is created.
    - Ties are identified based on multiple cars sharing the same Rank value
      within the top 3 positions.
    """
    try:
        df_all = read_excel_sheet(filepath)

        if "Racers" not in df_all:
            print("[INFO] No 'Racers' sheet to update.")
            return

        racers_df = df_all["Racers"]
        racers_df["Car"] = racers_df["Car"].astype(str)
        runoff_rows = []

        for sheet_name, df in df_all.items():
            if not sheet_name.endswith("_Rankings"):
                continue

            if "Rank" not in df.columns or "Car" not in df.columns:
                continue

            df["Car"] = df["Car"].astype(str)
            top3_df = df[df["Rank"].isin([1, 2, 3])].copy()
            rank_counts = top3_df["Rank"].value_counts()

            if any(rank_counts.get(r, 0) != 1 for r in [1, 2, 3]):
                enriched = top3_df.merge(
                    racers_df[["Car", "Class", "Group", "Description"]],
                    how="left",
                    on="Car",
                )
                for _, row in enriched.iterrows():
                    runoff_rows.append(
                        {
                            "Car": row["Car"],
                            "Name": row["Name"],
                            "Class": row["Class"],
                            "Group": row["Group"],
                            "Description": row["Description"],
                            "LastRank": row["Rank"],
                        }
                    )

        if not runoff_rows:
            print("[OK] No runoffs detected. No tab created.")
            return

        runoff_df = pd.DataFrame(runoff_rows)[
            ["Car", "Name", "Class", "Group", "Description", "LastRank"]
        ]
        runoff_df = runoff_df.sort_values(by=["Class", "Group", "LastRank", "Name"])

        book = load_workbook(filepath)
        if "Runoff" in book.sheetnames:
            book.remove(book["Runoff"])
        sheet = book.create_sheet("Runoff", index=1)

        for r_idx, row in enumerate(
            dataframe_to_rows(runoff_df, index=False, header=True), 1
        ):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

        book.save(filepath)
        print(f"[OK] Runoff tab created with {len(runoff_df)} entries.")

    except BaseException as err:
        print(f"[ERROR] Failed to create runoff tab: {err}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(f"Usage: python {sys.argv[0]} <results_file.xlsx>")
        sys.exit(1)

    arg_file = sys.argv[1]
    process_results(arg_file)
    add_runoff_tab(arg_file)
    format_all_sheets(arg_file)
