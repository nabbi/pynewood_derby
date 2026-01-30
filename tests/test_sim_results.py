"""Tests for sim_results.py — result simulation."""

import sys

import pandas as pd

from sim_results import main, simulate_and_write_results


class TestSimulateAndWriteResults:
    def test_fills_place_column(self, heats_xlsx):
        """After simulation, Place column should have valid integer placements."""
        # First, clear the Place column
        from openpyxl import load_workbook

        wb = load_workbook(str(heats_xlsx))
        ws = wb["Tiger_A"]
        # Find Place column
        place_col = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == "Place":
                place_col = idx
        assert place_col is not None
        # Clear placements
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=place_col).value = None
        wb.save(str(heats_xlsx))

        sheets = ["Tiger_A"]
        simulate_and_write_results(str(heats_xlsx), sheets)

        df = pd.read_excel(str(heats_xlsx), sheet_name="Tiger_A")
        assert df["Place"].notna().all()

    def test_placements_are_valid_within_heat(self, heats_xlsx):
        """Each heat's placements should be a permutation of 1..N."""
        simulate_and_write_results(str(heats_xlsx), ["Tiger_A"])
        df = pd.read_excel(str(heats_xlsx), sheet_name="Tiger_A")
        for _, group in df.groupby("Heat"):
            places = sorted(group["Place"].tolist())
            expected = list(range(1, len(group) + 1))
            assert places == expected

    def test_skips_racers_sheet(self, heats_xlsx):
        """Racers sheet should not be modified."""
        before = pd.read_excel(str(heats_xlsx), sheet_name="Racers")
        simulate_and_write_results(str(heats_xlsx), ["Racers", "Tiger_A"])
        after = pd.read_excel(str(heats_xlsx), sheet_name="Racers")
        pd.testing.assert_frame_equal(before, after)


class TestSimResultsMain:
    def test_main_runs(self, heats_xlsx, monkeypatch):
        monkeypatch.setattr(
            sys, "argv", ["sim_results.py", str(heats_xlsx)]
        )
        main()
        df = pd.read_excel(str(heats_xlsx), sheet_name="Tiger_A")
        assert df["Place"].notna().all()
